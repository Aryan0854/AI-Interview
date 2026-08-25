"""Import Resource_Question_Mapping.xlsx into employee portal tests."""
import base64
import hashlib
import json
import os
import re
import sys
import io
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
MAPPING_FILE = ROOT / "docs" / "Question Banks" / "Resource_Question_Mapping.xlsx"
QUESTION_BANK_FILE = ROOT / "docs" / "Question Banks" / "Question Bank-20th July '26.xlsx"
QB_SOURCE_FILE = ROOT / "docs" / "Question Banks" / "QB-new.xlsx"
ACCOUNTS_FILE = ROOT / "src" / "data" / "employee-accounts.json"
LOCAL_TESTS_FILE = ROOT / "uploads" / "local_tests_db.json"
MANIFEST_FILE = ROOT / "uploads" / "employee_test_manifest.json"

TOPIC_ID = "resource-product-assessment"
SUBJECT_ID = "resource-subject"
SUBJECT_TITLE = "Product Assessment"
TIME_LIMIT_SECONDS = 1800


def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()


def normalize_text(text: str) -> str:
    text = clean_str(text).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def dedupe_questions(questions: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for question in questions:
        q = clean_str(question)
        if not q:
            continue
        key = normalize_text(q)
        if key in seen:
            continue
        seen.add(key)
        unique.append(q)
    return unique


def strip_category_prefix(text: str) -> str:
    return re.sub(r"^\[[^\]]+\]\s*", "", clean_str(text))


def load_question_bank():
    """Load MCQ bank from generated Question Bank export."""
    bank: dict[str, dict] = {}

    def flush_current(current, current_key):
        if not current or len(current["options"]) < 2:
            return
        key = normalize_text(current["question_text"])
        if key not in bank:
            bank[key] = current.copy()

    # Try generated question bank export first
    if QUESTION_BANK_FILE.exists():
        wb = openpyxl.load_workbook(QUESTION_BANK_FILE, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header = [clean_str(c) for c in rows[0]]
            col = {h.lower(): i for i, h in enumerate(header)}

            q_col = next((col[k] for k in col if k == "question" or k.endswith("question")), None)
            opt_col = next((col[k] for k in col if "option" in k), None)
            correct_col = next((col[k] for k in col if "correct" in k), None)
            prod_col = next((col[k] for k in col if "product" in k), None)
            cat_col = next((col[k] for k in col if "category" in k), None)

            if q_col is None or opt_col is None:
                continue

            current_key = None
            current = None

            for row in rows[1:]:
                q_text = clean_str(row[q_col]) if q_col < len(row) else ""
                option = clean_str(row[opt_col]) if opt_col is not None and opt_col < len(row) else ""
                if not q_text or not option:
                    continue

                key = normalize_text(q_text)
                if key != current_key:
                    flush_current(current, current_key)
                    current_key = key
                    product = clean_str(row[prod_col]) if prod_col is not None and prod_col < len(row) else sheet_name
                    category = clean_str(row[cat_col]) if cat_col is not None and cat_col < len(row) else ""
                    plain = q_text
                    display = f"[{category}] {plain}" if category else plain
                    current = {
                        "question_text": display,
                        "plain_text": plain,
                        "options": [],
                        "correct_option_index": 0,
                        "explanation": "Imported from Question Bank.",
                        "difficulty": "medium",
                        "product": product or sheet_name,
                        "category": category,
                        "topic_title": category or product or sheet_name,
                    }

                current["options"].append(option)
                if correct_col is not None and correct_col < len(row):
                    correctness = clean_str(row[correct_col]).lower()
                    if correctness in {"correct", "true", "yes"}:
                        current["correct_option_index"] = len(current["options"]) - 1

            flush_current(current, current_key)

    # Index by display text and plain text for mapping lookup
    alias_bank: dict[str, dict] = {}
    for item in bank.values():
        alias_bank[normalize_text(item["question_text"])] = item
        plain = item.get("plain_text") or strip_category_prefix(item["question_text"])
        alias_bank[normalize_text(plain)] = item

    return alias_bank


def load_mapping_rows():
    wb = openpyxl.load_workbook(MAPPING_FILE, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    header = [clean_str(c) for c in rows[0]]
    col = {h: i for i, h in enumerate(header)}

    question_cols = [col[f"Assigned Question {i}"] for i in range(1, 26) if f"Assigned Question {i}" in col]

    employees = []
    for row in rows[1:]:
        emp_id = clean_str(row[col["Emp ID"]])
        if not emp_id:
            continue
        assigned = dedupe_questions(
            [clean_str(row[i]) for i in question_cols if i < len(row) and clean_str(row[i])]
        )
        employees.append(
            {
                "employee_id": emp_id,
                "full_name": clean_str(row[col.get("Emp Name", 1)]),
                "email": clean_str(row[col.get("Nokia Email ID", 5)]),
                "role": clean_str(row[col.get("Role", 2)]) or "employee",
                "department": clean_str(row[col.get("Domain", 3)]) or clean_str(row[col.get("Product", 4)]) or "SDM",
                "product": clean_str(row[col.get("Product", 4)]),
                "remarks": clean_str(row[col.get("Remarks", len(header) - 1)]),
                "assigned_questions": assigned,
            }
        )
    return employees


def hash_password(password: str):
    salt = base64.b64encode(os.urandom(16)).decode("ascii")
    digest = hashlib.pbkdf2_hmac("sha512", password.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return base64.b64encode(digest).decode("ascii"), salt


def generate_password(full_name: str, employee_id: str) -> str:
    first = (full_name or "User").strip().split()[0]
    first = re.sub(r"[^a-zA-Z]", "", first) or "User"
    first = first[:1].upper() + first[1:].lower()
    digits = re.sub(r"\D", "", employee_id)
    last4 = digits[-4:] if len(digits) >= 4 else digits.zfill(4)
    return f"EMP@{first}{last4}"


def ensure_accounts(employees):
    store = {"employees": []}
    if ACCOUNTS_FILE.exists():
        store = json.loads(ACCOUNTS_FILE.read_text(encoding="utf-8"))

    account_map = {
        clean_str(acc.get("employee_id")).upper(): acc
        for acc in store.get("employees", [])
        if acc.get("employee_id")
    }

    created = 0
    for emp in employees:
        norm_id = emp["employee_id"].upper()
        existing = account_map.get(norm_id)
        if existing and existing.get("password_hash") and existing.get("password_salt"):
            continue

        password = generate_password(emp["full_name"], emp["employee_id"])
        password_hash, password_salt = hash_password(password)
        account_map[norm_id] = {
            **(existing or {}),
            "employee_id": emp["employee_id"],
            "full_name": emp["full_name"] or emp["employee_id"],
            "email": emp["email"] or (existing or {}).get("email", ""),
            "department": emp["department"],
            "role": emp["role"] or (existing or {}).get("role", "employee"),
            "is_first_login": (existing or {}).get("is_first_login", True),
            "password_hash": password_hash,
            "password_salt": password_salt,
            "xp_points": (existing or {}).get("xp_points", 0),
            "streak_days": (existing or {}).get("streak_days", 0),
            "skill_level": (existing or {}).get("skill_level", "beginner"),
            "ai_readiness_score": (existing or {}).get("ai_readiness_score", 0),
        }
        created += 1

    ACCOUNTS_FILE.write_text(
        json.dumps({"employees": list(account_map.values())}, indent=2),
        encoding="utf-8",
    )
    return created


def build_tests(employees, bank):
    now = datetime.now(timezone.utc).isoformat()
    tests = []
    test_questions = []
    manifest = []
    stats = {
        "employees_total": len(employees),
        "employees_with_questions": 0,
        "employees_skipped": 0,
        "questions_matched": 0,
        "questions_missing": 0,
    }

    for emp in employees:
        if not emp["assigned_questions"]:
            stats["employees_skipped"] += 1
            continue

        matched = []
        missing = []
        for q_text in emp["assigned_questions"]:
            lookup = normalize_text(strip_category_prefix(q_text))
            item = bank.get(normalize_text(q_text)) or bank.get(lookup)
            if item:
                matched.append(
                    {
                        **item,
                        "question_text": q_text,
                    }
                )
                stats["questions_matched"] += 1
            else:
                missing.append(q_text)
                stats["questions_missing"] += 1

        if not matched:
            stats["employees_skipped"] += 1
            continue

        stats["employees_with_questions"] += 1
        test_id = str(uuid.uuid4())
        tests.append(
            {
                "id": test_id,
                "employee_id": emp["employee_id"],
                "topic_id": TOPIC_ID,
                "subject_id": SUBJECT_ID,
                "difficulty": "medium",
                "total_questions": len(matched),
                "time_limit_seconds": TIME_LIMIT_SECONDS,
                "status": "pending",
                "current_question_index": 0,
                "started_at": None,
                "completed_at": None,
                "in_progress": None,
                "created_at": now,
                "topic_title": emp["product"] or "Product Assessment",
                "subject_title": SUBJECT_TITLE,
            }
        )

        for idx, q in enumerate(matched):
            test_questions.append(
                {
                    "id": str(uuid.uuid4()),
                    "test_id": test_id,
                    "question_index": idx,
                    "question_text": q["question_text"],
                    "options": q["options"],
                    "correct_option_index": q["correct_option_index"],
                    "explanation": q["explanation"],
                    "difficulty": q["difficulty"],
                    "topic_id": TOPIC_ID,
                    "topic_title": q.get("category") or q["topic_title"],
                    "created_at": now,
                }
            )

        manifest.append(
            {
                "employee_id": emp["employee_id"],
                "full_name": emp["full_name"],
                "product": emp["product"],
                "test_id": test_id,
                "question_count": len(matched),
                "missing_questions": missing,
            }
        )

    return tests, test_questions, manifest, stats


def main():
    print("Loading Question Bank...")
    bank = load_question_bank()
    print(f"Loaded {len(bank)} unique MCQ questions from Question Bank.")

    print("Loading Resource Question Mapping...")
    employees = load_mapping_rows()
    print(f"Loaded {len(employees)} employees from mapping file.")

    created_accounts = ensure_accounts(employees)
    print(f"Ensured employee accounts ({created_accounts} new accounts added).")

    tests, test_questions, manifest, stats = build_tests(employees, bank)

    LOCAL_TESTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    LOCAL_TESTS_FILE.write_text(
        json.dumps(
            {"tests": tests, "test_questions": test_questions, "test_attempts": []},
            indent=2,
        ),
        encoding="utf-8",
    )

    MANIFEST_FILE.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    print("\n=== IMPORT SUMMARY ===")
    print(f"Employees total:            {stats['employees_total']}")
    print(f"Employees with tests:       {stats['employees_with_questions']}")
    print(f"Employees skipped:          {stats['employees_skipped']}")
    print(f"Questions matched:          {stats['questions_matched']}")
    print(f"Questions missing from QB:  {stats['questions_missing']}")
    print(f"Tests created:              {len(tests)}")
    print(f"Test questions created:     {len(test_questions)}")
    print(f"Saved local tests DB:       {LOCAL_TESTS_FILE}")
    print(f"Saved manifest:             {MANIFEST_FILE}")


if __name__ == "__main__":
    main()
