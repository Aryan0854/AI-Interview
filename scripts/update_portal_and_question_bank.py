"""
Update Employee Portal mapping from resources spreadsheet, then assign questions from QB-new.xlsx.

Step 1: Merge employee records into Resource_Question_Mapping.xlsx
Step 2: Build Question Bank from QB-new.xlsx and assign 25 questions per employee product
        using product-specific random / stratified sampling rules.
"""
from __future__ import annotations

import io
import json
import random
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from qb_new_parser import (
    QUESTIONS_PER_EMPLOYEE,
    assign_display_questions,
    mcq_to_export_rows,
    parse_qb_new_xlsx,
    resolve_qb_product_key,
)

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
RESOURCES_FILE = ROOT / "resources less than 3.5 rating - latest.xlsx"
QB_FILE = ROOT / "QB-new.xlsx"
MAPPING_FILE = ROOT / "Resource_Question_Mapping.xlsx"
QUESTION_BANK_FILE = ROOT / "Question Bank-20th July '26.xlsx"
LOG_FILE = ROOT / "uploads" / "portal_qb_update_log.json"

ASSIGNED_COLS = [f"Assigned Question {i}" for i in range(1, QUESTIONS_PER_EMPLOYEE + 1)]

OUTPUT_HEADER = [
    "Emp ID",
    "Emp Name",
    "Role",
    "Domain",
    "Product",
    "Nokia Email ID",
    "DDH",
    "Emp Status",
    *ASSIGNED_COLS,
    "Remarks",
]


def clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def normalize_question_key(question: str) -> str:
    text = clean(question)
    text = re.sub(r"^\[[^\]]+\]\s*", "", text)
    return re.sub(r"\s+", " ", text).lower()


def dedupe_questions(questions: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for question in questions:
        q = clean(question)
        if not q:
            continue
        key = normalize_question_key(q)
        if key in seen:
            continue
        seen.add(key)
        unique.append(q)
    return unique


def load_existing_mapping() -> dict[str, dict]:
    if not MAPPING_FILE.exists():
        return {}
    wb = openpyxl.load_workbook(MAPPING_FILE, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [clean(c) for c in rows[0]]
    col = {h: i for i, h in enumerate(header)}
    existing: dict[str, dict] = {}
    for row in rows[1:]:
        emp_id = clean(row[col["Emp ID"]] if "Emp ID" in col else "")
        if not emp_id:
            continue
        assigned = [
            clean(row[col[c]])
            for c in ASSIGNED_COLS
            if c in col and col[c] < len(row) and clean(row[col[c]])
        ]
        existing[emp_id] = {
            "emp_id": emp_id,
            "emp_name": clean(row[col.get("Emp Name", 1)] if col.get("Emp Name") is not None else ""),
            "role": clean(row[col.get("Role", 2)] if col.get("Role") is not None else ""),
            "domain": clean(row[col.get("Domain", 3)] if col.get("Domain") is not None else ""),
            "product": clean(row[col.get("Product", 4)] if col.get("Product") is not None else ""),
            "email": clean(row[col.get("Nokia Email ID", 5)] if col.get("Nokia Email ID") is not None else ""),
            "ddh": clean(row[col.get("DDH", 6)] if col.get("DDH") is not None else ""),
            "emp_status": clean(row[col.get("Emp Status", 7)] if col.get("Emp Status") is not None else ""),
            "assigned_questions": assigned,
            "remarks": clean(row[col.get("Remarks", len(header) - 1)] if "Remarks" in col else ""),
        }
    return existing


def load_resources_employees() -> list[dict]:
    wb = openpyxl.load_workbook(RESOURCES_FILE, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"No rows found in {RESOURCES_FILE.name}")
    header = [clean(c) for c in rows[0]]
    col = {h: i for i, h in enumerate(header)}

    def get(row, key, default=""):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return default
        return clean(row[idx])

    employees = []
    for row in rows[1:]:
        emp_id = get(row, "Emp ID")
        if not emp_id:
            continue
        employees.append(
            {
                "emp_id": emp_id,
                "emp_name": get(row, "Emp Name"),
                "email": get(row, "Nokia Email ID"),
                "role": get(row, "Role"),
                "domain": get(row, "Domain"),
                "product": get(row, "Product-Updated") or get(row, "Product"),
                "ddh": get(row, "DDH"),
            }
        )
    return employees


def get_question_pool(emp_product: str, pools: dict) -> tuple[list, str | None]:
    qb_key = resolve_qb_product_key(emp_product)
    if not qb_key:
        return [], f"No question bank mapping for product '{emp_product}'."

    if qb_key in pools and pools[qb_key]:
        return pools[qb_key], None

    parts = [p.strip() for p in re.split(r"[/]", emp_product) if p.strip()]
    combined = []
    for part in parts:
        pk = resolve_qb_product_key(part)
        if pk and pk in pools:
            combined.extend(pools[pk])
    if combined:
        return combined, None

    return [], f"Product '{emp_product}' not found in {QB_FILE.name}."


def write_question_bank(mcq_records) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Question Bank"
    ws.append(["Domain", "Product", "Category", "Question", "Option", "Correct"])

    rows_written = 0
    for row in mcq_to_export_rows(mcq_records):
        ws.append(row)
        rows_written += 1

    wb.save(QUESTION_BANK_FILE)
    return rows_written


def write_mapping(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Question Mapping"
    ws.append(OUTPUT_HEADER)
    for row in rows:
        ws.append(
            [
                row["emp_id"],
                row["emp_name"],
                row["role"],
                row["domain"],
                row["product"],
                row["email"],
                row["ddh"],
                row["emp_status"],
                *row["assigned_questions"],
                row["remarks"],
            ]
        )
    wb.save(MAPPING_FILE)


def validate_output(rows: list[dict], expected_count: int) -> list[str]:
    issues: list[str] = []
    if len(rows) != expected_count:
        issues.append(f"Row count mismatch: expected {expected_count}, got {len(rows)}")

    ids = [r["emp_id"] for r in rows]
    if len(ids) != len(set(ids)):
        issues.append("Duplicate Emp ID values detected in output.")

    for row in rows:
        qs = dedupe_questions([q for q in row["assigned_questions"] if q])
        if len(qs) != len([q for q in row["assigned_questions"] if q]):
            issues.append(f"Duplicate assigned questions for Emp ID {row['emp_id']}.")
        elif len(qs) != len(set(normalize_question_key(q) for q in qs)):
            issues.append(f"Duplicate assigned questions for Emp ID {row['emp_id']}.")

    return issues


def main() -> int:
    random.seed(42)
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)

    log = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "resources_file": RESOURCES_FILE.name,
        "qb_file": QB_FILE.name,
        "portal_updated": 0,
        "qb_records_written": 0,
        "question_assignments_updated": 0,
        "skipped": [],
        "validation_issues": [],
    }

    if not RESOURCES_FILE.exists():
        print(f"ERROR: Missing {RESOURCES_FILE.name}")
        return 1
    if not QB_FILE.exists():
        print(f"ERROR: Missing {QB_FILE.name}")
        return 1

    print("=== Step 1: Employee Portal update ===")
    existing = load_existing_mapping()
    resources_employees = load_resources_employees()
    print(f"Loaded {len(resources_employees)} employees from resources file.")
    if existing:
        print(f"Found {len(existing)} existing mapping records (fields will merge, questions reassigned in step 2).")

    merged_rows: list[dict] = []
    portal_updated = 0
    skipped_portal: list[dict] = []

    for emp in resources_employees:
        emp_id = emp["emp_id"]
        if not emp["emp_name"] or not emp["email"]:
            skipped_portal.append({"emp_id": emp_id, "reason": "Missing required name or email"})
            continue

        prior = existing.get(emp_id, {})
        merged = {
            "emp_id": emp_id,
            "emp_name": emp["emp_name"] or prior.get("emp_name", ""),
            "role": emp["role"] or prior.get("role", ""),
            "domain": emp["domain"] or prior.get("domain", ""),
            "product": emp["product"] or prior.get("product", ""),
            "email": emp["email"] or prior.get("email", ""),
            "ddh": emp["ddh"] or prior.get("ddh", ""),
            "emp_status": prior.get("emp_status", "Confirmed"),
            "assigned_questions": prior.get("assigned_questions", []),
            "remarks": prior.get("remarks", ""),
        }
        merged_rows.append(merged)
        portal_updated += 1

    log["portal_updated"] = portal_updated
    log["skipped"].extend(skipped_portal)

    print("=== Step 2: Question Bank + assignments (QB-new.xlsx) ===")
    pools, mcq_records = parse_qb_new_xlsx(QB_FILE)
    qb_rows = write_question_bank(mcq_records)
    log["qb_records_written"] = qb_rows
    print(f"Wrote {qb_rows} option rows to {QUESTION_BANK_FILE.name}.")
    print(f"QB products: {', '.join(f'{k}({len(v)})' for k, v in sorted(pools.items()))}")

    assignments_updated = 0
    assignment_failures: list[dict] = []

    for row in merged_rows:
        pool, pool_err = get_question_pool(row["product"], pools)
        if pool_err:
            row["assigned_questions"] = [""] * QUESTIONS_PER_EMPLOYEE
            row["remarks"] = pool_err
            assignment_failures.append({"emp_id": row["emp_id"], "product": row["product"], "reason": pool_err})
            continue

        qb_key = resolve_qb_product_key(row["product"]) or row["product"]
        assigned, remark = assign_display_questions(
            qb_key,
            pool,
            employee_id=row["emp_id"],
        )
        row["assigned_questions"] = assigned
        if remark:
            row["remarks"] = remark
        elif row["remarks"] and "not found" in row["remarks"].lower():
            row["remarks"] = ""
        assignments_updated += 1

    log["question_assignments_updated"] = assignments_updated
    log["skipped"].extend(assignment_failures)

    validation_issues = validate_output(merged_rows, len(resources_employees))
    log["validation_issues"] = validation_issues
    if validation_issues:
        print("VALIDATION ISSUES:")
        for issue in validation_issues:
            print(f"  - {issue}")

    write_mapping(merged_rows)
    LOG_FILE.write_text(json.dumps(log, indent=2), encoding="utf-8")

    print("\n=== SUMMARY ===")
    print(f"Employee Portal records updated: {portal_updated}")
    print(f"Question Bank option rows written: {qb_rows}")
    print(f"Question assignments updated: {assignments_updated}")
    print(f"Skipped / failed records: {len(log['skipped'])}")
    if log["skipped"]:
        for item in log["skipped"][:20]:
            print(f"  - Emp {item.get('emp_id')}: {item.get('reason')}")
        if len(log["skipped"]) > 20:
            print(f"  ... and {len(log['skipped']) - 20} more (see {LOG_FILE.name})")
    print(f"Output: {MAPPING_FILE.name}")
    print(f"Log: {LOG_FILE.name}")

    return 1 if validation_issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
