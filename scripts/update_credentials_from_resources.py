"""Build Employee_User_Credentials.xlsx from resources less than 3.5 rating - latest.xlsx only."""
from __future__ import annotations

import io
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
EXCEL_DIR = ROOT / "excel"
CREDENTIALS_FILE = EXCEL_DIR / "Employee_User_Credentials.xlsx"
RESOURCES_FILE = EXCEL_DIR / "resources less than 3.5 rating - latest.xlsx"

HEADER = [
    "Emp ID",
    "Employee Name",
    "Initial Password",
    "Nokia Email ID",
    "Role",
    "Domain",
    "Product",
    "Customer",
    "DDH Manager",
]


def clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def normalize_id(val) -> str:
    return clean(val).upper()


def generate_password(full_name: str, employee_id: str) -> str:
    first = (full_name or "User").split()[0]
    first = re.sub(r"[^a-zA-Z]", "", first) or "User"
    first = first[:1].upper() + first[1:].lower()
    digits = re.sub(r"\D", "", employee_id)
    last4 = digits[-4:] if len(digits) >= 4 else digits.zfill(4)
    return f"EMP@{first}{last4}"


def load_existing_passwords() -> dict[str, str]:
    if not CREDENTIALS_FILE.exists():
        return {}

    wb = openpyxl.load_workbook(CREDENTIALS_FILE, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [clean(c) for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}
    passwords: dict[str, str] = {}

    for row in rows[1:]:
        emp_id = clean(row[col["Emp ID"]] if "Emp ID" in col else "")
        if not emp_id:
            continue
        passwords[normalize_id(emp_id)] = clean(row[col.get("Initial Password", 2)])

    return passwords


def load_resources() -> list[dict]:
    wb = openpyxl.load_workbook(RESOURCES_FILE, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    header = [clean(c) for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}

    def get(row, key, default=""):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return default
        return clean(row[idx])

    employees = []
    for row in rows[1:]:
        emp_id = get(row, "Emp ID")
        if not emp_id:
            continue
        employees.append(
            {
                "Emp ID": emp_id,
                "Employee Name": get(row, "Emp Name"),
                "Nokia Email ID": get(row, "Nokia Email ID"),
                "Role": get(row, "Role"),
                "Domain": get(row, "Domain"),
                "Product": get(row, "Product-Updated") or get(row, "Product"),
                "Customer": get(row, "Customer"),
                "DDH Manager": get(row, "DDH"),
            }
        )
    return employees


def style_workbook(ws) -> None:
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4F46E5")
        cell.alignment = openpyxl.styles.Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 28

    widths = [16, 28, 22, 34, 20, 16, 24, 18, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width


def main() -> int:
    if not RESOURCES_FILE.exists():
        print(f"ERROR: Missing {RESOURCES_FILE.name}")
        return 1

    existing_passwords = load_existing_passwords()
    resources = load_resources()

    output: list[dict] = []
    passwords_generated = 0
    passwords_preserved = 0
    skipped: list[dict] = []

    for src in resources:
        if not src["Employee Name"] or not src["Nokia Email ID"]:
            skipped.append({"emp_id": src["Emp ID"], "reason": "Missing name or email in resources"})
            continue

        norm_id = normalize_id(src["Emp ID"])
        password = existing_passwords.get(norm_id, "")
        if password:
            passwords_preserved += 1
        else:
            password = generate_password(src["Employee Name"], src["Emp ID"])
            passwords_generated += 1

        output.append(
            {
                "Emp ID": src["Emp ID"],
                "Employee Name": src["Employee Name"],
                "Initial Password": password,
                "Nokia Email ID": src["Nokia Email ID"],
                "Role": src["Role"],
                "Domain": src["Domain"],
                "Product": src["Product"],
                "Customer": src["Customer"],
                "DDH Manager": src["DDH Manager"],
            }
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Credentials"
    ws.append(HEADER)
    for entry in output:
        ws.append([entry[h] for h in HEADER])
        row = ws.max_row
        ws.cell(row=row, column=3).font = Font(name="Consolas", size=10, bold=True, color="4338CA")
        ws.cell(row=row, column=1).value = str(entry["Emp ID"])

    style_workbook(ws)
    wb.save(CREDENTIALS_FILE)

    print("=== Employee_User_Credentials.xlsx created ===")
    print(f"Source file:                   {RESOURCES_FILE.name}")
    print(f"Rows written (resources only): {len(output)}")
    print(f"Passwords preserved:           {passwords_preserved}")
    print(f"Passwords generated:           {passwords_generated}")
    print(f"Skipped / failed:              {len(skipped)}")
    if skipped:
        for item in skipped:
            print(f"  - Emp {item['emp_id']}: {item['reason']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
