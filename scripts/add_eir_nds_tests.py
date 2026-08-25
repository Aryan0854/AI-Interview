"""
Assign EIR/NDS questions from QB-new.xlsx for employees missing product tests,
then merge those tests into local_tests_db + manifest (uploads + src/data).
Does NOT regenerate tests for employees who already have one.
"""
from __future__ import annotations

import json
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from qb_new_parser import (  # noqa: E402
    QUESTIONS_PER_EMPLOYEE,
    assign_display_questions,
    normalize_question_key,
    parse_qb_new_xlsx,
)

ACCOUNTS_FILE = ROOT / "src" / "data" / "employee-accounts.json"
EXCEL_DIR = ROOT / "excel"
MAPPING_FILE = EXCEL_DIR / "Resource_Question_Mapping.xlsx"
QB_FILE = EXCEL_DIR / "QB-new.xlsx"
TOPIC_ID = "resource-product-assessment"
SUBJECT_ID = "resource-subject"
SUBJECT_TITLE = "Product Assessment"
TIME_LIMIT_SECONDS = 1800
TARGET_PRODUCTS = {"EIR", "NDS"}

LOCAL_DB_PATHS = [
    ROOT / "uploads" / "local_tests_db.json",
    ROOT / "src" / "data" / "local_tests_db.json",
]
MANIFEST_PATHS = [
    ROOT / "uploads" / "employee_test_manifest.json",
    ROOT / "src" / "data" / "employee_test_manifest.json",
]


def clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def load_accounts() -> dict:
    return json.loads(ACCOUNTS_FILE.read_text(encoding="utf-8"))


def load_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def existing_test_employee_ids() -> set[str]:
    ids: set[str] = set()
    for path in LOCAL_DB_PATHS + MANIFEST_PATHS:
        data = load_json(path, {})
        if isinstance(data, list):
            for row in data:
                emp = clean(row.get("employee_id"))
                if emp:
                    ids.add(emp.upper())
        else:
            for test in data.get("tests", []):
                emp = clean(test.get("employee_code") or test.get("employee_id"))
                if emp:
                    ids.add(emp.upper())
    return ids


def update_mapping_row(emp: dict, assigned: list[str]) -> None:
    wb = openpyxl.load_workbook(MAPPING_FILE)
    ws = wb.active
    header = [clean(c.value) for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header)}
    emp_col = col.get("Emp ID")
    if not emp_col:
        raise RuntimeError("Emp ID column missing in Resource_Question_Mapping.xlsx")

    target_row = None
    for r in range(2, ws.max_row + 1):
        if clean(ws.cell(r, emp_col).value).upper() == emp["employee_id"].upper():
            target_row = r
            break

    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(target_row, emp_col).value = emp["employee_id"]

    field_map = {
        "Emp Name": emp.get("full_name", ""),
        "Role": emp.get("role", "employee"),
        "Domain": emp.get("department", ""),
        "Product": emp.get("product", ""),
        "Nokia Email ID": emp.get("email", ""),
        "Emp Status": "Confirmed",
        "Remarks": "",
    }
    for key, value in field_map.items():
        if key in col:
            ws.cell(target_row, col[key]).value = value

    for i, q in enumerate(assigned[:QUESTIONS_PER_EMPLOYEE], start=1):
        key = f"Assigned Question {i}"
        if key in col:
            ws.cell(target_row, col[key]).value = q

    wb.save(MAPPING_FILE)


def build_bank_index(mcq_records) -> dict:
    bank = {}
    for rec in mcq_records:
        item = {
            "question_text": rec.display_text,
            "plain_text": rec.question_text,
            "options": list(rec.options),
            "correct_option_index": rec.correct_option_index,
            "explanation": "Imported from QB-new.xlsx.",
            "difficulty": "medium",
            "product": rec.product,
            "category": rec.category,
            "topic_title": rec.category or rec.product,
        }
        bank[normalize_question_key(rec.display_text)] = item
        bank[normalize_question_key(rec.question_text)] = item
    return bank


def main() -> int:
    if not QB_FILE.exists():
        print(f"ERROR: missing {QB_FILE}")
        return 1
    if not ACCOUNTS_FILE.exists():
        print(f"ERROR: missing {ACCOUNTS_FILE}")
        return 1

    accounts = load_accounts()
    already = existing_test_employee_ids()
    targets = [
        e
        for e in accounts.get("employees", [])
        if e.get("product_qb_eligible")
        and clean(e.get("product")).upper() in TARGET_PRODUCTS
        and clean(e.get("employee_id")).upper() not in already
    ]
    if not targets:
        print("No missing EIR/NDS employees found. Nothing to do.")
        return 0

    print(f"Missing EIR/NDS employees: {len(targets)}")
    for e in targets:
        print(f"  - {e['employee_id']} ({e.get('product')}) {e.get('full_name')}")

    pools, mcq_records = parse_qb_new_xlsx(QB_FILE)
    bank = build_bank_index(mcq_records)
    now = datetime.now(timezone.utc).isoformat()

    new_tests = []
    new_questions = []
    new_manifest = []

    for emp in targets:
        product = clean(emp.get("product")).upper()
        pool = pools.get(product, [])
        assigned, remark = assign_display_questions(product, pool, employee_id=emp["employee_id"])
        assigned = [q for q in assigned if q]
        if len(assigned) < QUESTIONS_PER_EMPLOYEE:
            print(f"ERROR: {emp['employee_id']} only got {len(assigned)} questions. Remark: {remark}")
            return 1

        matched = []
        missing = []
        for q_text in assigned:
            item = bank.get(normalize_question_key(q_text)) or bank.get(
                normalize_question_key(q_text.split("] ", 1)[-1])
            )
            if not item:
                missing.append(q_text)
                continue
            matched.append({**item, "question_text": q_text})

        if missing or len(matched) != QUESTIONS_PER_EMPLOYEE:
            print(f"ERROR: match failure for {emp['employee_id']}: matched={len(matched)} missing={len(missing)}")
            for m in missing[:5]:
                print(f"  missing: {m[:120]}")
            return 1

        update_mapping_row(emp, assigned)

        test_id = str(uuid.uuid4())
        new_tests.append(
            {
                "id": test_id,
                "employee_id": emp["employee_id"],
                "employee_code": emp["employee_id"],
                "topic_id": TOPIC_ID,
                "subject_id": SUBJECT_ID,
                "difficulty": "medium",
                "total_questions": len(matched),
                "time_limit_seconds": TIME_LIMIT_SECONDS,
                "status": "pending",
                "current_question_index": 0,
                "started_at": None,
                "completed_at": None,
                "in_progress": None,
                "created_at": now,
                "topic_title": product or "Product Assessment",
                "subject_title": SUBJECT_TITLE,
            }
        )
        for idx, q in enumerate(matched):
            new_questions.append(
                {
                    "id": str(uuid.uuid4()),
                    "test_id": test_id,
                    "question_index": idx,
                    "question_text": q["question_text"],
                    "options": q["options"],
                    "correct_option_index": q["correct_option_index"],
                    "explanation": q["explanation"],
                    "difficulty": q["difficulty"],
                    "topic_id": TOPIC_ID,
                    "topic_title": q.get("category") or q["topic_title"],
                    "created_at": now,
                }
            )
        new_manifest.append(
            {
                "employee_id": emp["employee_id"],
                "full_name": emp.get("full_name", ""),
                "product": product,
                "test_id": test_id,
                "question_count": len(matched),
                "missing_questions": [],
            }
        )
        print(f"Created test {test_id} for {emp['employee_id']} ({product})")

    for path in LOCAL_DB_PATHS:
        db = load_json(path, {"tests": [], "test_questions": [], "test_attempts": []})
        db.setdefault("tests", []).extend(new_tests)
        db.setdefault("test_questions", []).extend(new_questions)
        db.setdefault("test_attempts", [])
        save_json(path, db)
        print(f"Updated {path} -> tests={len(db['tests'])} questions={len(db['test_questions'])}")

    for path in MANIFEST_PATHS:
        manifest = load_json(path, [])
        if not isinstance(manifest, list):
            manifest = []
        manifest.extend(new_manifest)
        save_json(path, manifest)
        print(f"Updated {path} -> entries={len(manifest)}")

    print("\nDone.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
