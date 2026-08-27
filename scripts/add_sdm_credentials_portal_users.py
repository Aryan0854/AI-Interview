"""
Add people from SDM Crediantials to be created.xlsx to Employee Portal.
Does not change existing portal users' tests, scores, or passwords.

Outputs:
  - NON-Needed docs/SDM_Credentials_Created.xlsx (shareable logins)
  - NON-Needed docs/Employee_User_Credentials.xlsx (append new only)
  - NON-Needed docs/Resource_Question_Mapping.xlsx (append new only)
  - src/data employee accounts, profiles, manifest, local tests
  - Supabase employees + tests
"""
from __future__ import annotations

import base64
import hashlib
import json
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import requests

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from qb_new_parser import (  # noqa: E402
    QUESTIONS_PER_EMPLOYEE,
    assign_display_questions,
    normalize_question_key,
    parse_qb_new_xlsx,
    resolve_qb_product_key,
)
from reassign_test_questions_supabase import (  # noqa: E402
    TOPIC_ID,
    SupabaseClient,
    load_env,
)

DOCS = ROOT / "NON-Needed docs"
QB_FILE = DOCS / "QB-new.xlsx"
CREDENTIALS_FILE = DOCS / "Employee_User_Credentials.xlsx"
MAPPING_FILE = DOCS / "Resource_Question_Mapping.xlsx"
SHARE_FILE = DOCS / "SDM_Credentials_Created.xlsx"
ACCOUNTS_FILE = ROOT / "src" / "data" / "employee-accounts.json"
PROFILES_FILE = ROOT / "src" / "data" / "resource_portal_profiles.json"

SUBJECT_ID = "resource-subject"
SUBJECT_TITLE = "Product Assessment"
TIME_LIMIT_SECONDS = 1800
REMARKS = "SDM credentials created"

LOCAL_DB_PATHS = [
    ROOT / "uploads" / "local_tests_db.json",
    ROOT / "src" / "data" / "local_tests_db.json",
]
MANIFEST_PATHS = [
    ROOT / "uploads" / "employee_test_manifest.json",
    ROOT / "src" / "data" / "employee_test_manifest.json",
]

SHARE_HEADER = [
    "Emp ID",
    "Employee Name",
    "Initial Password",
    "Nokia Email ID",
    "Role",
    "Domain",
    "Product",
    "Market",
    "Infra",
    "Login ID",
    "Portal status",
]


def clean(val) -> str:
    if val is None:
        return ""
    return str(val).replace("\xa0", " ").strip()


def load_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def hash_password(password: str) -> tuple[str, str]:
    salt = base64.b64encode(os.urandom(16)).decode("ascii")
    digest = hashlib.pbkdf2_hmac("sha512", password.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return base64.b64encode(digest).decode("ascii"), salt


def first_name_token(full_name: str) -> str:
    parts = re.findall(r"[A-Za-z]+", full_name or "")
    for part in parts:
        if len(part) >= 2:
            return part[:1].upper() + part[1:].lower()
    return "User"


def generate_password(full_name: str, employee_id: str) -> str:
    first = first_name_token(full_name)
    digits = re.sub(r"\D", "", employee_id)
    last4 = digits[-4:] if len(digits) >= 4 else digits.zfill(4)
    password = f"EMP@{first}{last4}"
    if not (
        len(password) >= 8
        and re.search(r"[A-Z]", password)
        and re.search(r"[a-z]", password)
        and re.search(r"[0-9]", password)
        and re.search(r"[^A-Za-z0-9]", password)
    ):
        raise RuntimeError(f"Generated password does not meet complexity rules: {password}")
    return password


def is_placeholder(val: str) -> bool:
    return clean(val).upper() in {"", "#N/A", "N/A", "NA", "NONE", "NULL"}


def is_valid_emp_id(emp_id: str) -> bool:
    if not emp_id:
        return False
    upper = emp_id.upper()
    if upper in {"EMP ID", "#N/A", "N/A", "NA", "NONE"}:
        return False
    return bool(re.search(r"\d", emp_id))


def is_valid_email(email: str) -> bool:
    if is_placeholder(email):
        return False
    return "@" in email


def find_source_file() -> Path:
    exact = DOCS / "SDM Crediantials to be created.xlsx"
    if exact.exists():
        return exact
    matches = []
    for path in DOCS.glob("*.xlsx"):
        name = path.name.lower()
        if "sdm" not in name:
            continue
        if "crediant" in name or "credentials to be created" in name:
            matches.append(path)
    if len(matches) == 1:
        return matches[0]
    raise FileNotFoundError(
        "Missing 'NON-Needed docs/SDM Crediantials to be created.xlsx'. Put that workbook back and run again."
    )


def load_existing_passwords() -> dict[str, str]:
    if not CREDENTIALS_FILE.exists():
        return {}
    ws = openpyxl.load_workbook(CREDENTIALS_FILE, data_only=True).active
    header = [clean(c.value) for c in ws[1]]
    col = {h: i for i, h in enumerate(header)}
    emp_i = col.get("Emp ID", 0)
    pw_i = col.get("Initial Password")
    out = {}
    if pw_i is None:
        return out
    for row in ws.iter_rows(min_row=2, values_only=True):
        eid = clean(row[emp_i])
        pw = clean(row[pw_i])
        if eid and pw:
            out[eid.upper()] = pw
    return out


def load_source_employees(source_file: Path) -> list[dict]:
    raw = source_file.read_bytes()
    if raw[:2] == b"PK":
        rows = list(openpyxl.load_workbook(source_file, data_only=True).active.iter_rows(values_only=True))
    else:
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("cp1252")
        delimiter = "\t" if "\t" in text.splitlines()[0] else ","
        rows = [line.split(delimiter) for line in text.splitlines() if line.strip()]

    header = [clean(c) for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}

    def get(row, *keys):
        for key in keys:
            if key in col and col[key] < len(row):
                value = clean(row[col[key]])
                if value and not is_placeholder(value):
                    return value
        return ""

    employees = []
    seen = set()
    for row in rows[1:]:
        emp_id = get(row, "Emp ID")
        if not is_valid_emp_id(emp_id) or emp_id.upper() in seen:
            continue
        seen.add(emp_id.upper())
        raw_product = get(row, "Product")
        product = resolve_qb_product_key(raw_product) or raw_product
        employees.append(
            {
                "employee_id": emp_id,
                "full_name": get(row, "Emp Name"),
                "email": get(row, "Nokia Email", "Email", "Nokia Email ID"),
                "nokia_emp_id": get(row, "Nokia Emp Id", "Nokia Emp", "Nokia Emp ID"),
                "role": get(row, "Role") or "Solution Engineer",
                "domain": get(row, "Domain") or "SDM",
                "product": product,
                "raw_product": raw_product,
                "market": get(row, "Market"),
                "infra": get(row, "Infra"),
                "ddh": "",
            }
        )
    return employees


def already_in_portal(emp_id: str, account_map: dict, profile_ids: set[str]) -> bool:
    key = emp_id.upper()
    acc = account_map.get(key) or {}
    return bool(acc.get("product_qb_eligible")) or key in profile_ids


def append_credentials_rows(employees: list[dict]) -> None:
    if not employees:
        return
    if CREDENTIALS_FILE.exists():
        wb = openpyxl.load_workbook(CREDENTIALS_FILE)
        ws = wb.active
        header = [clean(c.value) for c in ws[1]]
        col = {h: i + 1 for i, h in enumerate(header)}
        emp_col = col.get("Emp ID", 1)
        existing = {
            clean(ws.cell(r, emp_col).value).upper()
            for r in range(2, ws.max_row + 1)
            if clean(ws.cell(r, emp_col).value)
        }
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Credentials"
        header = [
            "Emp ID",
            "Employee Name",
            "Initial Password",
            "Nokia Email ID",
            "Role",
            "Domain",
            "Product",
            "Customer",
            "DDH Manager",
        ]
        ws.append(header)
        col = {h: i + 1 for i, h in enumerate(header)}
        emp_col = 1
        existing = set()

    added = 0
    for emp in employees:
        if emp["employee_id"].upper() in existing:
            continue
        row = ws.max_row + 1
        values = {
            "Emp ID": emp["employee_id"],
            "Employee Name": emp["full_name"],
            "Initial Password": emp["password"],
            "Nokia Email ID": emp["email"],
            "Role": emp["role"],
            "Domain": emp["domain"],
            "Product": emp["product"],
            "Customer": emp.get("market") or "",
            "DDH Manager": emp["ddh"],
        }
        for key, value in values.items():
            if key in col:
                ws.cell(row, col[key]).value = value
        ws.cell(row, emp_col).value = str(emp["employee_id"])
        if "Initial Password" in col:
            ws.cell(row, col["Initial Password"]).font = Font(
                name="Consolas", size=10, bold=True, color="4338CA"
            )
        added += 1
        existing.add(emp["employee_id"].upper())
    wb.save(CREDENTIALS_FILE)
    print(f"Appended {added} rows to {CREDENTIALS_FILE.name}")


def write_share_file(employees: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SDM Credentials"
    ws.append(SHARE_HEADER)
    for emp in employees:
        ws.append(
            [
                emp["employee_id"],
                emp["full_name"],
                emp["password"],
                emp["email"],
                emp["role"],
                emp["domain"],
                emp["product"],
                emp.get("market") or "",
                emp.get("infra") or "",
                emp["employee_id"],
                emp["portal_status"],
            ]
        )
        row = ws.max_row
        ws.cell(row, 1).value = str(emp["employee_id"])
        ws.cell(row, 10).value = str(emp["employee_id"])
        ws.cell(row, 3).font = Font(name="Consolas", size=10, bold=True, color="4338CA")

    header_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 28
    widths = [14, 32, 22, 42, 20, 12, 12, 12, 10, 14, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    note = wb.create_sheet("How to login")
    note["A1"] = "Employee Portal login"
    note["A1"].font = Font(bold=True, size=14)
    note["A3"] = "URL"
    note["B3"] = "https://ai-interview-ics-poc.vercel.app/employee"
    note["A4"] = "Login ID"
    note["B4"] = "Emp ID from this sheet"
    note["A5"] = "Password"
    note["B5"] = "Initial Password column"
    note["A7"] = "Newly added = created now with 25 product questions. Already in portal = existing password/tests were not changed."
    note["A8"] = "TM has no question bank, so those people are listed here but were not given a test."
    note["A9"] = "Missing Nokia Email / #N/A rows were skipped."
    note["A10"] = "Treat this file as confidential."
    note.column_dimensions["A"].width = 16
    note.column_dimensions["B"].width = 110
    wb.save(SHARE_FILE)
    print(f"Wrote share file: {SHARE_FILE} ({len(employees)} users)")


def update_mapping_row(emp: dict, assigned: list[str]) -> None:
    if not MAPPING_FILE.exists():
        print(f"Skip mapping workbook (missing): {MAPPING_FILE.name}")
        return
    wb = openpyxl.load_workbook(MAPPING_FILE)
    ws = wb.active
    header = [clean(c.value) for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header)}
    emp_col = col.get("Emp ID")
    if not emp_col:
        raise RuntimeError("Emp ID column missing in Resource_Question_Mapping.xlsx")

    target_row = None
    for r in range(2, ws.max_row + 1):
        if clean(ws.cell(r, emp_col).value).upper() == emp["employee_id"].upper():
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(target_row, emp_col).value = str(emp["employee_id"])

    field_map = {
        "Emp Name": emp["full_name"],
        "Role": emp["role"],
        "Domain": emp["domain"],
        "Product": emp["product"],
        "Nokia Email ID": emp["email"],
        "DDH": emp["ddh"],
        "Emp Status": "Confirmed",
        "Remarks": REMARKS,
    }
    for key, value in field_map.items():
        if key in col:
            ws.cell(target_row, col[key]).value = value
    for i, q in enumerate(assigned[:QUESTIONS_PER_EMPLOYEE], start=1):
        key = f"Assigned Question {i}"
        if key in col:
            ws.cell(target_row, col[key]).value = q
    wb.save(MAPPING_FILE)


def build_bank_index(mcq_records) -> dict:
    bank = {}
    for rec in mcq_records:
        item = {
            "question_text": rec.display_text,
            "plain_text": rec.question_text,
            "options": list(rec.options),
            "correct_option_index": rec.correct_option_index,
            "explanation": "Imported from QB-new.xlsx.",
            "difficulty": "medium",
            "product": rec.product,
            "category": rec.category,
            "topic_title": rec.category or rec.product,
        }
        bank[normalize_question_key(rec.display_text)] = item
        bank[normalize_question_key(rec.question_text)] = item
    return bank


def append_local_tests(new_tests: list[dict], new_questions: list[dict], new_manifest: list[dict]) -> None:
    ids = {t["id"] for t in new_tests}

    for path in LOCAL_DB_PATHS:
        if not path.exists() and path.parent.name == "uploads":
            continue
        db = load_json(path, {"tests": [], "test_questions": [], "test_attempts": []})
        db["tests"] = [t for t in db.get("tests", []) if t.get("id") not in ids]
        keep_test_ids = {t.get("id") for t in db["tests"]}
        db["test_questions"] = [q for q in db.get("test_questions", []) if q.get("test_id") in keep_test_ids]
        db["tests"].extend(new_tests)
        db["test_questions"].extend(new_questions)
        db.setdefault("test_attempts", [])
        save_json(path, db)
        print(f"Updated {path} -> tests={len(db['tests'])}")

    emp_ids = {clean(t.get("employee_id")).upper() for t in new_tests}
    for path in MANIFEST_PATHS:
        if not path.exists() and path.parent.name == "uploads":
            continue
        manifest = load_json(path, [])
        if not isinstance(manifest, list):
            manifest = []
        manifest = [row for row in manifest if clean(row.get("employee_id")).upper() not in emp_ids]
        manifest.extend(new_manifest)
        save_json(path, manifest)
        print(f"Updated {path} -> entries={len(manifest)}")


def raise_for_status(resp, action: str) -> None:
    if resp.ok:
        return
    raise RuntimeError(f"{action} failed {resp.status_code}: {resp.text}")


def upload_storage_object(url: str, key: str, bucket: str, object_path: str, data: bytes, content_type: str) -> None:
    resp = requests.put(
        f"{url}/storage/v1/object/{bucket}/{object_path}",
        headers={
            "Authorization": f"Bearer {key}",
            "apikey": key,
            "Content-Type": content_type,
            "x-upsert": "true",
        },
        data=data,
        timeout=120,
    )
    raise_for_status(resp, f"upload {bucket}/{object_path}")


def persist_shared_files() -> None:
    url, key = load_env()
    if MAPPING_FILE.exists():
        upload_storage_object(
            url,
            key,
            "docs-ingest",
            "Portal Mapping/Resource_Question_Mapping.xlsx",
            MAPPING_FILE.read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        print("Uploaded mapping workbook to Portal Mapping cloud folder")
    manifest_path = ROOT / "src" / "data" / "employee_test_manifest.json"
    if manifest_path.exists():
        upload_storage_object(
            url,
            key,
            "app-data",
            "employee_test_manifest.json",
            manifest_path.read_bytes(),
            "application/json",
        )
        print("Uploaded employee_test_manifest.json to app-data")


def upsert_supabase_employee(client: SupabaseClient, emp: dict, account: dict) -> str:
    payload = {
        "employee_id": account["employee_id"],
        "email": account["email"],
        "full_name": account["full_name"],
        "department": "general",
        "role": account.get("role") or "employee",
        "product": account.get("product"),
        "is_first_login": True,
        "password_hash": account["password_hash"],
        "password_salt": account["password_salt"],
        "product_qb_eligible": True,
        "assessment_only": True,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    existing = client.get("employees", {"employee_id": f"eq.{emp['employee_id']}", "select": "id", "limit": "1"})
    if existing:
        resp = requests.patch(
            f"{client.base}/employees",
            headers=client.headers,
            params={"employee_id": f"eq.{emp['employee_id']}"},
            json=payload,
            timeout=60,
        )
        raise_for_status(resp, f"PATCH employees {emp['employee_id']}")
        return existing[0]["id"]
    resp = requests.post(f"{client.base}/employees", headers=client.headers, json=[payload], timeout=60)
    raise_for_status(resp, f"POST employees {emp['employee_id']}")
    created = client.get("employees", {"employee_id": f"eq.{emp['employee_id']}", "select": "id", "limit": "1"})
    if not created:
        raise RuntimeError(f"Failed to create Supabase employee {emp['employee_id']}")
    return created[0]["id"]


def create_test_if_missing(client: SupabaseClient, employee_uuid: str, emp: dict, matched: list[dict], test_id: str) -> None:
    existing = client.get(
        "tests",
        {
            "employee_id": f"eq.{employee_uuid}",
            "topic_id": f"eq.{TOPIC_ID}",
            "select": "id,status",
            "limit": "1",
        },
    )
    if existing:
        print(f"  skip existing test {existing[0]['id']} status={existing[0].get('status')}")
        return

    now = datetime.now(timezone.utc).isoformat()
    resp = requests.post(
        f"{client.base}/tests",
        headers=client.headers,
        json=[
            {
                "id": test_id,
                "employee_id": employee_uuid,
                "employee_code": emp["employee_id"],
                "topic_id": TOPIC_ID,
                "subject_id": SUBJECT_ID,
                "topic_title": emp["product"],
                "subject_title": SUBJECT_TITLE,
                "difficulty": "medium",
                "total_questions": len(matched),
                "time_limit_seconds": TIME_LIMIT_SECONDS,
                "status": "pending",
                "current_question_index": 0,
            }
        ],
        timeout=60,
    )
    raise_for_status(resp, f"POST tests {emp['employee_id']}")
    payload = []
    for idx, q in enumerate(matched):
        payload.append(
            {
                "id": str(uuid.uuid4()),
                "test_id": test_id,
                "question_index": idx,
                "question_text": q["question_text"],
                "options": q["options"],
                "correct_option_index": q["correct_option_index"],
                "explanation": q.get("explanation") or "Imported from QB-new.xlsx.",
                "difficulty": q.get("difficulty") or "medium",
                "topic_id": TOPIC_ID,
                "topic_title": q.get("category") or q.get("topic_title") or emp["product"],
                "created_at": now,
            }
        )
    for i in range(0, len(payload), 50):
        client.post("test_questions", payload[i : i + 50])


def main() -> int:
    try:
        source_file = find_source_file()
    except FileNotFoundError as err:
        print(f"ERROR: {err}")
        return 1
    if not QB_FILE.exists():
        print(f"ERROR: missing {QB_FILE}")
        return 1

    employees = load_source_employees(source_file)
    stored_passwords = load_existing_passwords()
    print(f"Loaded {len(employees)} employees from {source_file.name}")

    accounts = load_json(ACCOUNTS_FILE, {"employees": []})
    account_map = {
        clean(acc.get("employee_id")).upper(): acc for acc in accounts.get("employees", []) if acc.get("employee_id")
    }
    profiles = load_json(PROFILES_FILE, [])
    profile_map = {clean(row.get("employee_id")).upper(): row for row in profiles}
    profile_ids = set(profile_map)

    to_add = []
    share_rows = []
    skipped_no_qb = []
    skipped_missing = []
    for emp in employees:
        if not emp["full_name"] or not is_valid_email(emp["email"]):
            emp["password"] = "(not created — missing name or email)"
            emp["portal_status"] = "Not added: missing name or email"
            emp["skip"] = True
            skipped_missing.append(emp)
            share_rows.append(emp)
            continue

        if already_in_portal(emp["employee_id"], account_map, profile_ids):
            emp["password"] = stored_passwords.get(emp["employee_id"].upper()) or "(already in portal — password unchanged)"
            emp["portal_status"] = "Already in portal"
            emp["skip"] = True
            share_rows.append(emp)
            continue

        known = resolve_qb_product_key(emp["product"])
        pool_ok = known in {"HLR-HSS", "SDL", "UDM", "EIR", "NDS", "CMG", "CMM", "NRD"}
        if not pool_ok:
            emp["password"] = "(not created — no question bank for this product)"
            emp["portal_status"] = f"Not added: no question bank for {emp['raw_product'] or emp['product']}"
            emp["skip"] = True
            skipped_no_qb.append(emp)
            share_rows.append(emp)
            continue

        emp["password"] = stored_passwords.get(emp["employee_id"].upper()) or generate_password(
            emp["full_name"], emp["employee_id"]
        )
        emp["password_hash"], emp["password_salt"] = hash_password(emp["password"])
        emp["portal_status"] = "Newly added"
        emp["skip"] = False
        to_add.append(emp)
        share_rows.append(emp)

    print(f"Already in portal (unchanged): {sum(1 for e in share_rows if e.get('portal_status') == 'Already in portal')}")
    print(f"Missing name/email: {len(skipped_missing)}")
    print(f"No question bank: {len(skipped_no_qb)}")
    print(f"To add: {len(to_add)}")

    pools, mcq_records = parse_qb_new_xlsx(QB_FILE)
    bank = build_bank_index(mcq_records)
    now = datetime.now(timezone.utc).isoformat()
    new_tests = []
    new_questions = []
    new_manifest = []

    for emp in to_add:
        product = emp["product"]
        pool = pools.get(product, [])
        assigned, remark = assign_display_questions(product, pool, employee_id=emp["employee_id"])
        assigned = [q for q in assigned if q]
        if len(assigned) < QUESTIONS_PER_EMPLOYEE:
            print(f"ERROR: {emp['employee_id']} only got {len(assigned)} {product} questions. {remark}")
            return 1

        matched = []
        for q_text in assigned:
            item = bank.get(normalize_question_key(q_text)) or bank.get(
                normalize_question_key(q_text.split("] ", 1)[-1])
            )
            if not item:
                print(f"ERROR: unmatched question for {emp['employee_id']}: {q_text[:120]}")
                return 1
            matched.append({**item, "question_text": q_text})

        existing_acc = account_map.get(emp["employee_id"].upper(), {})
        account = {
            **existing_acc,
            "employee_id": emp["employee_id"],
            "full_name": emp["full_name"],
            "email": emp["email"],
            "department": emp["domain"] or "SDM",
            "role": emp["role"],
            "product": product,
            "assessment_only": True,
            "product_qb_eligible": True,
            "is_first_login": True,
            "password_hash": emp["password_hash"],
            "password_salt": emp["password_salt"],
            "xp_points": existing_acc.get("xp_points", 0),
            "streak_days": existing_acc.get("streak_days", 0),
            "skill_level": existing_acc.get("skill_level", "beginner"),
            "ai_readiness_score": existing_acc.get("ai_readiness_score", 0),
        }
        account_map[emp["employee_id"].upper()] = account
        emp["account"] = account
        emp["assigned"] = assigned
        emp["matched"] = matched

        profile_map[emp["employee_id"].upper()] = {
            "employee_id": emp["employee_id"],
            "full_name": emp["full_name"],
            "role": emp["role"],
            "domain": emp["domain"],
            "product": product,
            "email": emp["email"],
            "ddh": emp["ddh"],
            "emp_status": "Confirmed",
            "remarks": REMARKS,
            "assigned_questions": assigned,
            "assigned_question_count": len(assigned),
        }

        update_mapping_row(emp, assigned)

        test_id = str(uuid.uuid4())
        emp["test_id"] = test_id
        new_tests.append(
            {
                "id": test_id,
                "employee_id": emp["employee_id"],
                "employee_code": emp["employee_id"],
                "topic_id": TOPIC_ID,
                "subject_id": SUBJECT_ID,
                "difficulty": "medium",
                "total_questions": len(matched),
                "time_limit_seconds": TIME_LIMIT_SECONDS,
                "status": "pending",
                "current_question_index": 0,
                "started_at": None,
                "completed_at": None,
                "in_progress": None,
                "created_at": now,
                "topic_title": product,
                "subject_title": SUBJECT_TITLE,
            }
        )
        for idx, q in enumerate(matched):
            new_questions.append(
                {
                    "id": str(uuid.uuid4()),
                    "test_id": test_id,
                    "question_index": idx,
                    "question_text": q["question_text"],
                    "options": q["options"],
                    "correct_option_index": q["correct_option_index"],
                    "explanation": q["explanation"],
                    "difficulty": q["difficulty"],
                    "topic_id": TOPIC_ID,
                    "topic_title": q.get("category") or q["topic_title"],
                    "created_at": now,
                }
            )
        new_manifest.append(
            {
                "employee_id": emp["employee_id"],
                "full_name": emp["full_name"],
                "product": product,
                "test_id": test_id,
                "question_count": len(matched),
                "missing_questions": [],
            }
        )
        print(f"Prepared {emp['employee_id']} {emp['full_name']} -> {product} ({len(matched)} Qs)")

    accounts["employees"] = list(account_map.values())
    save_json(ACCOUNTS_FILE, accounts)
    save_json(PROFILES_FILE, list(profile_map.values()))
    append_credentials_rows(to_add)
    write_share_file(share_rows)
    append_local_tests(new_tests, new_questions, new_manifest)

    print(f"Syncing {len(to_add)} new portal users to Supabase...")
    url, key = load_env()
    client = SupabaseClient(url, key)
    for emp in to_add:
        employee_uuid = upsert_supabase_employee(client, emp, emp["account"])
        create_test_if_missing(client, employee_uuid, emp, emp["matched"], emp["test_id"])
        print(f"  synced {emp['employee_id']}")

    persist_shared_files()

    print("\n=== SDM credentials portal add complete ===")
    print(f"Unchanged existing portal users: {sum(1 for e in share_rows if e.get('portal_status') == 'Already in portal')}")
    print(f"Skipped (missing name/email): {len(skipped_missing)}")
    print(f"Skipped (no question bank): {len(skipped_no_qb)}")
    print(f"Newly added: {len(to_add)}")
    print(f"Share file: {SHARE_FILE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
