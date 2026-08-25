"""
Remap one employee from their current product to a new product (e.g. CMG -> CMM).
Updates local roster/mapping files and live Supabase test questions.

Usage:
  python scripts/remap_employee_product.py --employee 1042323 --product CMM
"""
from __future__ import annotations

import argparse
import json
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from qb_new_parser import (  # noqa: E402
    QUESTIONS_PER_EMPLOYEE,
    assign_display_questions,
    normalize_question_key,
    parse_qb_new_xlsx,
    resolve_qb_product_key,
)
from reassign_test_questions_supabase import (  # noqa: E402
    TOPIC_ID,
    SupabaseClient,
    load_env,
    reassign_one,
)

ACCOUNTS_FILE = ROOT / "src" / "data" / "employee-accounts.json"
PROFILES_FILE = ROOT / "src" / "data" / "resource_portal_profiles.json"
MANIFEST_FILE = ROOT / "src" / "data" / "employee_test_manifest.json"
LOCAL_TESTS_FILE = ROOT / "src" / "data" / "local_tests_db.json"
MAPPING_FILE = ROOT / "docs" / "Question Banks" / "Resource_Question_Mapping.xlsx"
QB_FILE = ROOT / "docs" / "Question Banks" / "QB-new.xlsx"


def clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def build_bank_index(mcq_records) -> dict:
    bank = {}
    for rec in mcq_records:
        item = {
            "question_text": rec.display_text,
            "options": list(rec.options),
            "correct_option_index": rec.correct_option_index,
            "explanation": "Imported from QB-new.xlsx.",
            "difficulty": "medium",
            "product": rec.product,
            "category": rec.category,
            "topic_title": rec.category or rec.product,
        }
        bank[normalize_question_key(rec.display_text)] = item
        bank[normalize_question_key(rec.question_text)] = item
    return bank


def match_questions(assigned: list[str], bank: dict) -> list[dict]:
    matched = []
    missing = []
    for q_text in assigned:
        item = bank.get(normalize_question_key(q_text)) or bank.get(
            normalize_question_key(q_text.split("] ", 1)[-1])
        )
        if not item:
            missing.append(q_text)
            continue
        matched.append({**item, "question_text": q_text})
    if missing:
        raise RuntimeError(f"Could not match {len(missing)} question(s): {missing[:3]}")
    return matched


def update_mapping_excel(emp_id: str, product: str, assigned: list[str]) -> None:
    wb = openpyxl.load_workbook(MAPPING_FILE)
    ws = wb.active
    header = [clean(c.value) for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header)}
    emp_col = col.get("Emp ID")
    if not emp_col:
        raise RuntimeError("Emp ID column missing in Resource_Question_Mapping.xlsx")

    target_row = None
    for r in range(2, ws.max_row + 1):
        if clean(ws.cell(r, emp_col).value).upper() == emp_id.upper():
            target_row = r
            break
    if target_row is None:
        raise RuntimeError(f"Employee {emp_id} not found in Resource_Question_Mapping.xlsx")

    if "Product" in col:
        ws.cell(target_row, col["Product"]).value = product
    for i, q in enumerate(assigned[:QUESTIONS_PER_EMPLOYEE], start=1):
        key = f"Assigned Question {i}"
        if key in col:
            ws.cell(target_row, col[key]).value = q
    wb.save(MAPPING_FILE)


def update_local_tests(test_id: str, product: str, matched: list[dict]) -> None:
    if not LOCAL_TESTS_FILE.exists():
        return
    db = load_json(LOCAL_TESTS_FILE)
    now = datetime.now(timezone.utc).isoformat()
    for test in db.get("tests", []):
        if test.get("id") == test_id:
            test["topic_title"] = product
            test["status"] = "pending"
            test["current_question_index"] = 0
            test["started_at"] = None
            test["completed_at"] = None
            test["in_progress"] = None
            test["total_questions"] = len(matched)
            break
    remaining = [q for q in db.get("test_questions", []) if q.get("test_id") != test_id]
    for idx, q in enumerate(matched):
        remaining.append(
            {
                "id": str(uuid.uuid4()),
                "test_id": test_id,
                "question_index": idx,
                "question_text": q["question_text"],
                "options": q["options"],
                "correct_option_index": q["correct_option_index"],
                "explanation": q.get("explanation") or "Imported from QB-new.xlsx.",
                "difficulty": q.get("difficulty") or "medium",
                "topic_id": TOPIC_ID,
                "topic_title": product,
                "created_at": now,
            }
        )
    db["test_questions"] = remaining
    if "test_attempts" in db:
        db["test_attempts"] = [a for a in db["test_attempts"] if a.get("test_id") != test_id]
    save_json(LOCAL_TESTS_FILE, db)


def main() -> int:
    parser = argparse.ArgumentParser(description="Remap one employee to a new product question bank.")
    parser.add_argument("--employee", required=True, help="Employee ID, e.g. 1042323")
    parser.add_argument("--product", required=True, help="Target product, e.g. CMM")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    emp_id = clean(args.employee)
    product_key = resolve_qb_product_key(args.product)
    if not product_key:
        raise SystemExit(f"Unknown product '{args.product}'")

    accounts = load_json(ACCOUNTS_FILE)
    account = next(
        (e for e in accounts.get("employees", []) if clean(e.get("employee_id")) == emp_id),
        None,
    )
    if not account:
        raise SystemExit(f"Employee {emp_id} not found in employee-accounts.json")

    previous = clean(account.get("product"))
    print(f"{emp_id} {account.get('full_name')} : {previous} -> {product_key}")

    pools, mcq_records = parse_qb_new_xlsx(QB_FILE)
    pool = pools.get(product_key, [])
    assigned, remark = assign_display_questions(product_key, pool, employee_id=emp_id)
    assigned = [q for q in assigned if q]
    if len(assigned) < QUESTIONS_PER_EMPLOYEE:
        raise SystemExit(f"Only assigned {len(assigned)} questions. {remark}")
    matched = match_questions(assigned, build_bank_index(mcq_records))
    print(f"Assigned {len(matched)} {product_key} questions. Sample: {matched[0]['question_text'][:90]}")

    if args.dry_run:
        print("Dry run only. No files or Supabase updates.")
        return 0

    account["product"] = product_key
    save_json(ACCOUNTS_FILE, accounts)
    print(f"Updated {ACCOUNTS_FILE.name}")

    profiles = load_json(PROFILES_FILE)
    for row in profiles:
        if clean(row.get("employee_id")) == emp_id:
            row["product"] = product_key
            row["assigned_questions"] = assigned
            row["assigned_question_count"] = len(assigned)
            break
    save_json(PROFILES_FILE, profiles)
    print(f"Updated {PROFILES_FILE.name}")

    manifest = load_json(MANIFEST_FILE)
    test_id = None
    for row in manifest:
        if clean(row.get("employee_id")) == emp_id:
            row["product"] = product_key
            row["question_count"] = len(matched)
            row["missing_questions"] = []
            test_id = row.get("test_id")
            break
    save_json(MANIFEST_FILE, manifest)
    print(f"Updated {MANIFEST_FILE.name} (test_id={test_id})")

    update_mapping_excel(emp_id, product_key, assigned)
    print(f"Updated {MAPPING_FILE.name}")

    if test_id:
        update_local_tests(test_id, product_key, matched)
        print(f"Updated {LOCAL_TESTS_FILE.name}")

    url, key = load_env()
    client = SupabaseClient(url, key)
    client.patch("employees", {"employee_id": f"eq.{emp_id}"}, {"product": product_key})
    print("Updated Supabase employees.product")

    result = reassign_one(
        client,
        employee_code=emp_id,
        matched_questions=matched,
        dry_run=False,
    )
    if result["status"] != "updated":
        print(f"Supabase reassign failed: {result}")
        return 1

    live_test_id = result.get("test_id")
    if live_test_id:
        client.patch("tests", {"id": f"eq.{live_test_id}"}, {"topic_title": product_key})
    print(f"Reassigned live test {live_test_id} to {product_key} and reset status to pending.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
